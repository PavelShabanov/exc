from datetime import datetime as dt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import (column_index_from_string,
                                 coordinate_from_string)
from openpyxl.chart import (
    ScatterChart,
    LineChart,
    Reference,
    Series,
)

# заполнить вручную
file_name_read = 'Приложение 1__(2).xlsx' #название файла с данными
file_name_write = 'Приложение 1__(2)_res.xlsx' #название файла с результатами
sheet_name = 'ВЛГ' #'ВЛГ' #название листа в файле
name_start_column = 'A' #даты начала работ
name_end_column = 'B' #даты конца работ
name_costs_column ='E' #стоимости работ
name_table_start_column = 'L' #столбец, где начинается заполнение графика
ind_start_row = 12 #строка с 1-й работой
ind_end_row = 48 #строка с последней работой

month_last = 17 # сколько месяцев прошло с начала проекта (сколько заполнено данных по месяцам по EV и AC)
ind_EV_row = 52 # программа считывает - здесь вписываем каждый месяц выполнение из КС-2
ind_AC_row = 53 # программа считывает - здесь вписываем каждый месяц оплаты заказчика из КС-3 + авансы
name_unit = 'K' # программой - столбец, куда будут вписывать названия вычисляемых параметров
ind_month_enum_row = 50 # программой - в эту строку впишутся номера месяцев
ind_sum_row = 55 # программой - строка, где суммируются плановые затраты по месяцам
short_version = False # заполнить только таблицу-график, к-ты внизу не вычислять

#далее - работает программа
units_arr = ['мес.', '', 'EV', 'AC', '', 'PV', 'PV sum', 'PV av', 'EV', 'EV sum', 'EV av', 'AC', 'AC sum', 'AC av', \
    '--- БЮДЖЕТ ПРОЕКТА ---', 'BAC', 'CV>0', 'CPI>1', 'EAC', 'ETC', 'VAC', 'PC $', '--- СРОКИ ПРОЕКТА ---', \
    'SV', 'SPI', 'TCPI', 'PCwork', 'FtW', 'PC']
starts = []
ends = []
work_in_month = []
costs = []
start_month = []
end_month = []
costs_per_month = []
PV_in_month = []
PV_sum = []

BAC = 0.0
PV_aver = []
month_enum = []
EV_in_month = []
EV_sum = []
EV_aver = []
AC_in_month = []
AC_sum = []
AC_aver = []
CV = []
CPI = []
EAC = []
ETC = []
VAC = []
PCdollar = []
SV = []
SPI = []
TCPI = []
PCwork = []
FtW = []
PC = []


ind_start_column = ind_end_column = ind_costs_column = ind_table_start_column = 0
start_date_project = end_date_project = 0


#функции
def calc_columns_index():
    global ind_start_column, ind_end_column, ind_costs_column, ind_table_start_column, ind_unit_column
    xy = coordinate_from_string(name_start_column+str(ind_start_row))
    ind_start_column = column_index_from_string(xy[0])
    xy = coordinate_from_string(name_end_column+str(ind_start_row))
    ind_end_column = column_index_from_string(xy[0])
    xy = coordinate_from_string(name_costs_column+str(ind_start_row))
    ind_costs_column = column_index_from_string(xy[0])
    xy = coordinate_from_string(name_table_start_column+str(ind_start_row))
    ind_table_start_column = column_index_from_string(xy[0])
    xy = coordinate_from_string(name_unit+str(ind_start_row))
    ind_unit_column = column_index_from_string(xy[0])

def time_delta_in_month(d1, d2):
    m1 = d1.month
    y1 = d1.year
    m2 = d2.month
    y2 = d2.year
    if y1 == y2:
        return m2-m1+1
    elif y2 > y1:
        return (y2-y1-1)*12 + (12-m1+1)+m2
    else:
        return -111

def fill_month(m, y):
    global start_month, end_month
    if m==1 or m==3 or m==5 or m==7 or m==8 or m==10 or m==12:
        start_month.append(dt(y, m, 1).date())
        end_month.append(dt(y, m, 31).date())
    elif m==2:
        start_month.append(dt(y, m, 1).date())
        if y%4 == 0: #високосный
            end_month.append(dt(y, m, 29).date())
        else:
            end_month.append(dt(y, m, 28).date())
    elif m==4 or m==6 or m==9 or m==11:
        start_month.append(dt(y, m, 1).date())
        end_month.append(dt(y, m, 30).date())

def fill_month2(st_date, count_month):
    global start_month, end_month
    month = st_date.month
    year = st_date.year
    for i in range(0, count_month):
        fill_month(month, year)
        if month == 12:
            month = 1
            year += 1
        else:
            month += 1

def find_start_end_project(st, end):
    global start_date_project, end_date_project
    st_ = st.copy()
    st_.sort()
    end_ = end.copy()
    end_.sort()
    start_date_project = st_[0]
    end_date_project = end_[-1]

def insert_graf():
    global count_month_in_project, sheet_graf, ind_table_start_column, ind_month_enum_row,\
        ind_sum_row, add_row
    chart = ScatterChart()
    chart.title = "График выполнения работ"
    chart.style = 2
    chart.x_axis.title = 'месяцы'
    chart.y_axis.title = 'проценты'
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = count_month_in_project
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = BAC
    xvalues = Reference(sheet_graf, min_col=ind_table_start_column, min_row=ind_month_enum_row, \
        max_col=ind_table_start_column+count_month_in_project-1, max_row=ind_month_enum_row)
    # PV sum
    values = Reference(sheet_graf, min_col=ind_table_start_column-1, min_row=ind_sum_row+1, \
        max_col=ind_table_start_column+count_month_in_project-1, max_row=ind_sum_row+1)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)
    # AC sum
    values = Reference(sheet_graf, min_col=ind_table_start_column-1, min_row=ind_sum_row+1+6, \
        max_col=ind_table_start_column+count_month_in_project-1, max_row=ind_sum_row+1+6)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)
    # EV sum
    values = Reference(sheet_graf, min_col=ind_table_start_column-1, min_row=ind_sum_row+1+3, \
        max_col=ind_table_start_column+count_month_in_project-1, max_row=ind_sum_row+1+3)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)
    # добавляем диаграмму
    add_row += 1
    sheet_graf.add_chart(chart, str(name_table_start_column)+str(ind_sum_row+len(units_arr)+1)) 


    line1 = chart.series[0]
    # цвет заливки линии графика
    line1.graphicalProperties.line.solidFill = "FF9900"
    # символ маркера для текущего значения
    line1.marker.symbol = "x"
    # цвет заливки маркера
    line1.marker.graphicalProperties.solidFill = "FF9900"
    line1.marker.graphicalProperties.line.solidFill = "FF9900"
    # заливаем линию между маркерами (не прозрачная)
    line1.graphicalProperties.line.noFill = False
    # делаем линию гладкой
    line1.smooth = True
    # ширина указывается в EMU
    line1.graphicalProperties.line.width = 100050

    # ЛИНИЯ С ДАННЫМИ ИЗ 2 СТОЛБЦА ДАННЫХ
    line2 = chart.series[1]
    # цвет заливки линии графика
    line2.graphicalProperties.line.solidFill = "00AAFF"
    # делаем линию пунктирной
    line2.graphicalProperties.line.dashStyle = "sysDot"
    # ширина указывается в EMU
    line2.graphicalProperties.line.width = 100050

    # ЛИНИЯ С ДАННЫМИ ИЗ 3 СТОЛБЦА ДАННЫХ
    line3 = chart.series[2]
    # цвет заливки линии графика
    line3.graphicalProperties.line.solidFill = "00FF66"
    # символ маркера для текущего значения
    line3.marker.symbol = "triangle"
    # покрасим маркер в другой цвет
    line3.marker.graphicalProperties.solidFill = "00FF66"
    line3.marker.graphicalProperties.line.solidFill = "00FF66"
    # делаем линию гладкой
    line3.graphicalProperties.line.dashStyle = "sysDash"
    # ширина указывается в EMU
    line3.graphicalProperties.line.width = 100050

#открыть файл с нужным листом
book_graf=openpyxl.open(file_name_read) #, data_only=True)
sheet_graf=book_graf[sheet_name]
#вычислить индексы столбцов, откуда считывать данные (чтобы не использовать буквы в названиях столбцов)
calc_columns_index()
#считать данные о
# - начале работ (каждой работы)
# - конце работ
# - сколько месяцев длиться работа (с расчетом в функции)
# - стоимости работы
# - стоимости работы в месяц, но с учетом, что 1-й и посл-й месяцы идут с расходом "/2", остальные месяцы с полной суммой
for i in range(ind_start_row, ind_end_row+1):
    ind_row = i-ind_start_row
    starts.append(dt.date(sheet_graf.cell(row=i, column=ind_start_column).value))
    ends.append(dt.date(sheet_graf.cell(row=i, column=ind_end_column).value))
    work_in_month.append(time_delta_in_month(starts[ind_row], ends[ind_row])) #append((ends[j]-starts[j]).days//27)
    costs.append(sheet_graf.cell(row=i, column=ind_costs_column).value)
    costs_per_month.append(costs[ind_row]/(work_in_month[ind_row]))#-1)) при -1 можно сделать затраты на 1й и последний месяц в два раза меньше, чем со 2го по предпоследний
    # находим общую стоимость работ BAC
    BAC += costs[ind_row]
#находим самую раннюю и подзнюю даты, это начало и конец проекта; а также продолжительность в месяцах всего проекта
find_start_end_project(starts, ends)
count_month_in_project = time_delta_in_month(start_date_project, end_date_project)
# заполняем даты для каждой ячейки в укрупненном графике
fill_month2(start_date_project, count_month_in_project)
#заполнить укрупненный график с суммами по месяцам
#begin_to_fill = False
#begin_j = -1
for i in range(ind_start_row, ind_end_row+1):
    ind_row = i-ind_start_row
    for j in range(ind_table_start_column, ind_table_start_column+count_month_in_project):
        ind_col = j-ind_table_start_column
        if ((starts[ind_row]>=start_month[ind_col] and starts[ind_row]<=end_month[ind_col]) or 
            (ends[ind_row]>=start_month[ind_col] and ends[ind_row]<=end_month[ind_col]) or
            (starts[ind_row]<=start_month[ind_col] and ends[ind_row]>=end_month[ind_col])):
            #if begin_to_fill == False:
            #    begin_to_fill = True
            #    begin_j = j
            #if j == begin_j or j == begin_j+work_in_month[ind_row]: #-1:
            sheet_graf.cell(row=i, column=j).value = costs_per_month[ind_row] #/2.0
            #else:
            #    sheet_graf.cell(row=i, column=j).value = costs_per_month[ind_row]
            sheet_graf.cell(row=i, column=j).fill = openpyxl.styles.PatternFill(start_color="0033CCFF", end_color="0033CCFF", fill_type = "solid")
        else:
            sheet_graf.cell(row=i, column=j).value = ""
            sheet_graf.cell(row=i, column=j).fill = openpyxl.styles.PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type = "solid")
    #begin_to_fill = False
    #begin_j = -1

# выполнять ли программу дальше, если короткая версия, то очищаем пространство под таблицей,
# сохраняем файл и выходим из программы
if short_version == True:
    # стереть все под таблицей
    # стереть номера месяцев
    for j in range(ind_table_start_column, ind_table_start_column+count_month_in_project):
        ind_col = j-ind_table_start_column
        sheet_graf.cell(row=ind_month_enum_row, column=j).value = ''
    # стереть параметры под таблицей и EV и AC
    for i in range(ind_month_enum_row, ind_sum_row+len(units_arr)-(ind_sum_row-ind_EV_row)):
        for j in range(ind_unit_column, ind_unit_column+count_month_in_project+2):
            sheet_graf.cell(row=i, column=j).value = ''
    book_graf.save(file_name_write)
    quit()

# вписывать в столбцец названия параметров, которые вычисляются
for i in range(ind_month_enum_row, ind_month_enum_row+len(units_arr)): # -2 - чтобы подписать и месяцы
    ind_row = i-ind_month_enum_row
    sheet_graf.cell(row=i, column=ind_unit_column).value = units_arr[ind_row]

#заполняем все-все-все по месяцам и накопительным итогом
for j in range(ind_table_start_column, ind_table_start_column+count_month_in_project):
    ind_col = j-ind_table_start_column
    # заполнить номера месяцев под таблицей
    month_enum.append(ind_col+1)
    sheet_graf.cell(row=ind_month_enum_row, column=j).value = month_enum[ind_col]
    # заполняем PV по месяцам
    PV_in_month.append(0)
    for i in range(ind_start_row, ind_end_row+1):
        if sheet_graf.cell(row=i, column=j).data_type == 'n':
            PV_in_month[ind_col] += sheet_graf.cell(row=i, column=j).value
    #заполняем по месяцам
    add_row = 0
    sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PV_in_month[ind_col]
    #заполняем накопительно PV
    add_row += 1
    if ind_col == 0:
        PV_sum.append(PV_in_month[ind_col])
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PV_sum[ind_col]
    else:
        PV_sum.append(0)
        PV_sum[ind_col] = PV_sum[ind_col-1] + PV_in_month[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PV_sum[ind_col]
    # заполняем среднее PV
    add_row += 1
    PV_aver.append(0)
    PV_aver[ind_col] += PV_sum[ind_col]/month_enum[ind_col]
    sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PV_aver[ind_col]

    if ind_col<month_last:
        # заполняем EV
        # считываем
        EV_in_month.append(0)
        EV_in_month[ind_col] += sheet_graf.cell(row=ind_EV_row, column=j).value
        # вписываем ниже то же самое EV по месяцам
        add_row += 1
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = EV_in_month[ind_col]
        #заполняем накопительно EV
        add_row += 1
        if ind_col == 0:
            EV_sum.append(EV_in_month[ind_col])
            sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = EV_sum[ind_col]
        else:
            EV_sum.append(0)
            EV_sum[ind_col] = EV_sum[ind_col-1] + EV_in_month[ind_col]
            sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = EV_sum[ind_col]
        # заполняем среднее EV
        add_row += 1
        EV_aver.append(0)
        EV_aver[ind_col] += EV_sum[ind_col]/month_enum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = EV_aver[ind_col]
        # заполняем AC
        # считываем
        AC_in_month.append(0)
        AC_in_month[ind_col] += sheet_graf.cell(row=ind_AC_row, column=j).value
        # вписываем ниже то же самое AC по месяцам
        add_row += 1
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = AC_in_month[ind_col]
        #заполняем накопительно AC
        add_row += 1
        if ind_col == 0:
            AC_sum.append(AC_in_month[ind_col])
            sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = AC_sum[ind_col]
        else:
            AC_sum.append(0)
            AC_sum[ind_col] = AC_sum[ind_col-1] + AC_in_month[ind_col]
            sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = AC_sum[ind_col]
        # заполняем среднее AC
        add_row += 1
        AC_aver.append(0)
        AC_aver[ind_col] += AC_sum[ind_col]/month_enum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = AC_aver[ind_col]
        # пропуск строки
        add_row += 1
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = ''
        # заполняем BAC
        add_row += 1
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = BAC
        # заполняем CV
        add_row += 1
        CV.append(0)
        CV[ind_col] = EV_sum[ind_col] - AC_sum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = CV[ind_col]
        # заполняем CPI
        add_row += 1
        CPI.append(0.0)
        CPI[ind_col] = EV_sum[ind_col]/AC_sum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = CPI[ind_col]    
        # заполняем EAC
        add_row += 1
        EAC.append(0.0)
        EAC[ind_col] = BAC/CPI[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = EAC[ind_col]  
        # заполняем ETC
        add_row += 1
        ETC.append(0.0)
        ETC[ind_col] = EAC[ind_col] - AC_sum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = ETC[ind_col]  
        # заполняем VAC
        add_row += 1
        VAC.append(0.0)
        VAC[ind_col] = EAC[ind_col] - BAC
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = VAC[ind_col]  
        # заполняем PCdollar
        add_row += 1
        PCdollar.append(0.0)
        PCdollar[ind_col] = AC_sum[ind_col] / BAC
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PCdollar[ind_col]  
        # пропуск строки
        add_row += 1
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = ''
        # заполняем SV
        add_row += 1
        SV.append(0.0)
        SV[ind_col] = EV_sum[ind_col] - PV_sum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = SV[ind_col]  
        # заполняем SPI
        add_row += 1
        SPI.append(0.0)
        SPI[ind_col] = EV_sum[ind_col] / PV_sum[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = SPI[ind_col]  
        # заполняем TCPI
        add_row += 1
        TCPI.append(0.0)
        TCPI[ind_col] = (BAC - EV_sum[ind_col]) / (BAC - PV_sum[ind_col])
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = TCPI[ind_col]  
        # заполняем PCwork
        add_row += 1
        PCwork.append(0.0)
        PCwork[ind_col] = EV_sum[ind_col] / BAC
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PCwork[ind_col]  
        # заполняем FtW
        add_row += 1
        FtW.append(0.0)
        FtW[ind_col] = count_month_in_project / SPI[ind_col]
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = FtW[ind_col]  
        # заполняем PC
        add_row += 1
        PC.append(0.0)
        PC[ind_col] = EV_sum[ind_col] / BAC
        sheet_graf.cell(row=ind_sum_row+add_row, column=j).value = PC[ind_col]  

# строим диаграмму
insert_graf()

# сохранить файл
book_graf.save(file_name_write)